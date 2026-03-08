from flask import Flask, request, jsonify
from flask_cors import CORS
import os, hashlib, datetime
from enum import Enum
import pandas as pd
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_XLSX = os.path.join(BASE_DIR, "case_master.xlsx")


# FSM (backstage mechanics)
class FSMState(Enum):
    PLIE = "IDLE"
    GLISSADE = "EVALUATE"
    JETE = "COMMIT"
    FAULT = "HALT"


class BalanchineFSM:
    def __init__(self):
        self.state = FSMState.PLIE

    def next(self, s):
        self.state = s
        return self.state


def mercy_gate(conf: float = 0.0) -> bool:
    return conf >= 0.80


def receipt_hash(action: str, entity: str = "") -> str:
    payload = f"{action}|{entity}|{datetime.datetime.utcnow().isoformat()}"
    return hashlib.sha256(payload.encode()).hexdigest()


# Init ledger
if not os.path.exists(MASTER_XLSX):
    wb = Workbook()
    ws = wb.active
    ws.title = "Intake_Log"
    ws.append(["doc_id", "filename", "source", "url", "ingest_dt", "hash", "status"])
    wb.save(MASTER_XLSX)


@app.route('/')
def home():
    return "Prima Ledger active. Use /api/wings/ingest or /api/grid/execute"


@app.route('/api/wings/ingest', methods=['POST'])
def ingest():
    urls = request.json.get('urls', [])
    wb = load_workbook(MASTER_XLSX)
    ws = wb["Intake_Log"]
    added = 0

    for url in urls:
        h = hashlib.sha256(url.encode()).hexdigest()[:8].upper()
        doc_id = f"D-{h}"
        if not any(doc_id == r[0].value for r in ws.iter_rows(min_row=2)):
            ws.append([
                doc_id,
                url.split('/')[-1],
                "dataset",
                url,
                datetime.datetime.utcnow().isoformat(),
                h,
                "pending",
            ])
            added += 1

    wb.save(MASTER_XLSX)
    return jsonify({"added": added})


@app.route('/api/grid/execute', methods=['POST'])
def execute():
    d = request.json
    action = d.get('action')

    if action not in ('promote', 'discard'):
        return jsonify({"error": "Invalid action"}), 400

    fsm = BalanchineFSM()
    fsm.next(FSMState.GLISSADE)

    # Placeholder gate (expand with real entity lookup)
    if action == "promote" and not mercy_gate():
        fsm.next(FSMState.FAULT)
        receipt = receipt_hash("ATTEMPT_PROMOTE")
        return jsonify({"status": "HALTED", "receipt": receipt}), 403

    fsm.next(FSMState.JETE)
    receipt = receipt_hash(action.upper())
    return jsonify({"status": "COMMITTED", "receipt": receipt, "fsm": fsm.state.value})


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
