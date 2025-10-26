"""Generate the Field Coherence Dashboard Excel template.

This script mirrors the specification provided in the user prompt. It builds
the workbook with the following sheets:

* ``Data_Input`` – empty table ready for data entry.
* ``Config`` – min/max bounds, desired direction and weights.
* ``README`` – textual instructions for using the dashboard.
* ``Journal`` – lightweight reflective journaling template.
* ``Normalized`` – formulas that transform raw data into 0..100 scores.
* ``Dashboard`` – quick summary of the latest Field Coherence readings.

Running the script will write ``Field_Coherence_Dashboard_Template.xlsx`` into
the repository root (``/workspace/gemini-cli`` inside the execution
environment).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


OUTPUT_PATH = Path("Field_Coherence_Dashboard_Template.xlsx")


def build_base_workbook(path: Path) -> None:
    """Create the initial workbook populated with simple dataframes."""

    data_columns = [
        "Date",
        # Physiological
        "HRV (ms)",
        "Resting HR (bpm)",
        "Sleep (hrs)",
        "Body Temp Δ (°C)",
        # Cognitive/Behavioral
        "Screen Time (hrs)",
        "Typing Variance (ms)",
        "Calendar Switches (count)",
        # Relational
        "Msgs Sent (count)",
        "Msgs Received (count)",
        "Avg Sentiment (-1..1)",
        # Financial (optional)
        "Transactions (count)",
    ]
    data_df = pd.DataFrame(columns=data_columns)

    config_df = pd.DataFrame(
        {
            "Metric": [
                "HRV (ms)",
                "Resting HR (bpm)",
                "Sleep (hrs)",
                "Body Temp Δ (°C)",
                "Screen Time (hrs)",
                "Typing Variance (ms)",
                "Calendar Switches (count)",
                "Msgs Sent (count)",
                "Msgs Received (count)",
                "Avg Sentiment (-1..1)",
                "Transactions (count)",
            ],
            "Desired Direction": [
                "high_good",
                "low_good",
                "high_good",
                "low_good",
                "low_good",
                "low_good",
                "low_good",
                "balanced",
                "balanced",
                "high_good",
                "balanced",
            ],
            "Min": [20, 50, 4, 0.0, 1, 20, 5, 0, 0, -1, 0],
            "Max": [120, 85, 9, 1.0, 10, 200, 30, 200, 200, 1, 100],
            "Weight (0..1)": [
                0.18,
                0.10,
                0.12,
                0.08,
                0.10,
                0.07,
                0.07,
                0.07,
                0.07,
                0.10,
                0.04,
            ],
        }
    )

    readme_lines = [
        "FIELD COHERENCE DASHBOARD — Instructions",
        "",
        "Sheets:",
        "1) Data_Input — paste daily or weekly values. Add one row per date.",
        "2) Config — adjust Min/Max and Weights to match your physiology and habits.",
        "3) Normalized — auto-computed 0..100 scores per metric + domain indices.",
        "4) Dashboard — overall Field Coherence Index (0..100) and summaries.",
        "5) Journal — short reflective notes to add narrative texture to data.",
        "",
        "Normalization logic:",
        "For 'high_good' metrics: score = 100 * (value - Min) / (Max - Min).",
        "For 'low_good' metrics:  score = 100 * (Max - value) / (Max - Min).",
        "For 'balanced' metrics:  treated as high_good by default; adjust with Weights if needed.",
        "All scores are clipped to [0, 100].",
        "",
        "Domain groupings:",
        "- Physiological: HRV, Resting HR, Sleep, Body Temp Δ",
        "- Cognitive/Behavioral: Screen Time, Typing Variance, Calendar Switches",
        "- Relational/Network: Msgs Sent, Msgs Received, Avg Sentiment, Transactions",
        "",
        "Interpretation tips:",
        "- Divergence between domain indices suggests de-synchronization.",
        "- Sustained red (<40) is a cue to downshift workload and increase recovery.",
        "- Track weekly trends rather than obsess over single-day noise.",
    ]
    readme_df = pd.DataFrame({"Instructions": readme_lines})

    journal_df = pd.DataFrame(
        {
            "Date": [],
            "Body felt like…": [],
            "Mind behaved like…": [],
            "People around me felt like…": [],
            "Notes / Interventions": [],
        }
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        data_df.to_excel(writer, sheet_name="Data_Input", index=False)
        config_df.to_excel(writer, sheet_name="Config", index=False)
        readme_df.to_excel(writer, sheet_name="README", index=False)
        journal_df.to_excel(writer, sheet_name="Journal", index=False)


def add_formula_sheets(path: Path) -> None:
    """Append the Normalized and Dashboard sheets using openpyxl."""

    wb = load_workbook(path)

    config_sheet = wb["Config"]
    metrics = [cell.value for cell in config_sheet["A"][1:] if cell.value]
    weights_start_row = 2
    weights_end_row = weights_start_row + len(metrics) - 1

    # Normalized sheet -----------------------------------------------------
    ws_norm = wb.create_sheet("Normalized")
    headers = [
        "Date",
        *[f"{metric} (Score)" for metric in metrics],
        "Physiological Index",
        "Cognitive Index",
        "Relational Index",
        "Field Coherence Index",
    ]
    ws_norm.append(headers)

    # Mapping from metric -> column index in Data_Input (B..)
    metric_to_data_col = {metric: idx + 2 for idx, metric in enumerate(metrics)}
    config_row_for_metric = {metric: idx + 2 for idx, metric in enumerate(metrics)}

    for row in range(2, 1002):
        # Date column mirrors Data_Input
        ws_norm.cell(row=row, column=1).value = f"=Data_Input!A{row}"

        for metric_index, metric in enumerate(metrics):
            target_col = metric_index + 2  # B onwards
            data_col_letter = get_column_letter(metric_to_data_col[metric])
            config_row = config_row_for_metric[metric]

            min_cell = f"Config!C{config_row}"
            max_cell = f"Config!D{config_row}"
            direction_cell = f"Config!B{config_row}"
            value_cell = f"Data_Input!{data_col_letter}{row}"

            formula = (
                "=IFERROR(MAX(0,MIN(100,"
                f"IF({direction_cell}=\"low_good\","
                f"100*(({max_cell}-{value_cell})/({max_cell}-{min_cell})),"
                f"100*((({value_cell}-{min_cell}))/({max_cell}-{min_cell}))"
                "))),\"\")"
            )
            ws_norm.cell(row=row, column=target_col).value = formula

        # Domain indices
        physi_cols = [get_column_letter(col) for col in range(2, 6)]
        ws_norm.cell(row=row, column=len(headers) - 3).value = (
            "=IFERROR(AVERAGE("
            + ",".join(f"{col}{row}" for col in physi_cols)
            + "),\"\")"
        )

        cognitive_cols = [get_column_letter(col) for col in range(6, 9)]
        ws_norm.cell(row=row, column=len(headers) - 2).value = (
            "=IFERROR(AVERAGE("
            + ",".join(f"{col}{row}" for col in cognitive_cols)
            + "),\"\")"
        )

        relational_cols = [get_column_letter(col) for col in range(9, 13)]
        ws_norm.cell(row=row, column=len(headers) - 1).value = (
            "=IFERROR(AVERAGE("
            + ",".join(f"{col}{row}" for col in relational_cols)
            + "),\"\")"
        )

        score_start_col = get_column_letter(2)
        score_end_col = get_column_letter(1 + len(metrics))
        ws_norm.cell(row=row, column=len(headers)).value = (
            "=IFERROR("
            f"SUMPRODUCT({score_start_col}{row}:{score_end_col}{row},"
            f"Config!$E${weights_start_row}:$E${weights_end_row})/"
            f"SUM(Config!$E${weights_start_row}:$E${weights_end_row}),\"\")"
        )

    # Dashboard sheet ------------------------------------------------------
    ws_dash = wb.create_sheet("Dashboard")
    ws_dash["A1"] = "FIELD COHERENCE DASHBOARD"
    ws_dash["A3"] = "Latest Date"
    ws_dash["B3"] = "=MAX(Normalized!A2:A1001)"

    physi_col_letter = get_column_letter(len(headers) - 3)
    cognitive_col_letter = get_column_letter(len(headers) - 2)
    relational_col_letter = get_column_letter(len(headers) - 1)
    fci_col_letter = get_column_letter(len(headers))

    def latest_value_formula(column_letter: str) -> str:
        return (
            "=LET(d,Normalized!A2:A1001,"
            "r,MATCH(MAX(d),d,0),"
            f"INDEX(Normalized!{column_letter}2:{column_letter}1001,r))"
        )

    ws_dash["A5"] = "Physiological Index (Latest)"
    ws_dash["B5"] = latest_value_formula(physi_col_letter)

    ws_dash["A6"] = "Cognitive Index (Latest)"
    ws_dash["B6"] = latest_value_formula(cognitive_col_letter)

    ws_dash["A7"] = "Relational Index (Latest)"
    ws_dash["B7"] = latest_value_formula(relational_col_letter)

    ws_dash["A9"] = "Field Coherence Index (Latest)"
    ws_dash["B9"] = latest_value_formula(fci_col_letter)

    wb.save(path)


def main() -> None:
    build_base_workbook(OUTPUT_PATH)
    add_formula_sheets(OUTPUT_PATH)


if __name__ == "__main__":
    main()
